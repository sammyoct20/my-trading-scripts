import requests
import pandas as pd
from datetime import datetime
import os

# ---- CONFIG ----------------------------------------------------------
# Configured to run cleanly inside relative cloud repository root folder
OUTPUT_FOLDER = "."
OUTPUT_FILENAME = f"NSE_FnO_Top_Losers_{datetime.now().strftime('%Y-%m-%d_%H%M')}.xlsx"
TOP_N = 20
# -----------------------------------------------------------------------

class NSEFnOTopLosers:
    def __init__(self):
        self.session = requests.Session()
        self.headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                           "AppleWebKit/537.36 (KHTML, like Gecko) "
                           "Chrome/120.0.0.0 Safari/537.36",
            "Accept": "application/json, text/plain, */*",
            "Accept-Language": "en-US,en;q=0.9",
            "Referer": "https://www.nseindia.com/market-data/top-gainers-losers",
        }
        self.session.get("https://www.nseindia.com", headers=self.headers, timeout=5)
        self.session.get(
            "https://www.nseindia.com/market-data/top-gainers-losers",
            headers=self.headers, timeout=5
        )

    def get_fno_securities_losers(self, top_n=20):
        url = "https://www.nseindia.com/api/live-analysis-variations?index=loosers"
        response = self.session.get(url, headers=self.headers, timeout=10)
        response.raise_for_status()
        data = response.json()

        rows = data["FOSec"]["data"]
        timestamp = data["FOSec"]["timestamp"]

        df = pd.DataFrame(rows)
        df = df.head(top_n)

        cols = ["symbol", "open_price", "high_price", "low_price",
                "prev_price", "ltp", "perChange", "trade_quantity"]
        df = df[[c for c in cols if c in df.columns]]

        rename_map = {
            "symbol": "Symbol",
            "open_price": "Open",
            "high_price": "High",
            "low_price": "Low",
            "prev_price": "Prev. Close",
            "ltp": "LTP",
            "perChange": "% Change",
            "trade_quantity": "Volume (shares)",
        }
        df = df.rename(columns=rename_map)

        return df, timestamp


def save_to_excel(df, timestamp, folder, filename):
    os.makedirs(folder, exist_ok=True)
    full_path = os.path.join(folder, filename)

    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="FnO Top Losers", index=False)

        sheet = writer.sheets["FnO Top Losers"]

        from openpyxl.styles import Font, PatternFill, Alignment

        header_font = Font(name="Arial", bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")

        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        for col_cells in sheet.columns:
            length = max(len(str(cell.value)) for cell in col_cells if cell.value is not None)
            sheet.column_dimensions[col_cells[0].column_letter].width = max(12, length + 2)

        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")

        note_row = sheet.max_row + 2
        sheet.cell(row=note_row, column=1,
                    value=f"Source: NSE India (F&O Securities) | As of: {timestamp}")

    return full_path


if __name__ == "__main__":
    nse = NSEFnOTopLosers()
    df, ts = nse.get_fno_securities_losers(top_n=TOP_N)
    saved_path = save_to_excel(df, ts, OUTPUT_FOLDER, OUTPUT_FILENAME)
    print(f"Saved: {saved_path}")
    print(df.to_string(index=False))